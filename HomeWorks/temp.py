from openpyxl import load_workbook
import matplotlib.pyplot as plt


wb = load_workbook(r"C:\Users\ihor.ivchenko\PycharmProjects\python_courses\file.xlsx")

# help(wb)

sheet = wb.active

# print(sheet.max_column)
# print(sheet.max_row)

x = []
y = []

for row in range(1, sheet.max_row + 1):
    x.append(sheet.cell(row=row, column=1).value)
    y.append(sheet.cell(row=row, column=2).value)

plt.plot(x, y)
plt.show()